import openpyxl
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

class Pretty:
    # 直接复制请求头时以字符串形式转为字典格式
    def Pretty_headers(self,start_headers:str) -> dict:
        headers = {}
        start_headers = start_headers.strip().split("\n")
        for i in start_headers:
            head = i.strip().split(":",1)
            headers[head[0]] = head[1].strip()
        return headers

    # 优化Excel文件格式，使行宽列宽适合。
    def Pretty_excel(self,filename,sheetname) -> None:
        wb = openpyxl.load_workbook(filename)
        ws = wb[sheetname]

        df=pd.read_excel(filename,sheetname).fillna('-')
        df.loc[len(df)]=list(df.columns)						#把标题行附加到最后一行
        for col in df.columns:				
            index=list(df.columns).index(col)					#列序号
            letter=get_column_letter(index+1)					#列字母
            collen=df[col].apply(lambda x:len(str(x).encode())).max()	#获取这一列长度的最大值 也可以用min获取最小值 mean获取平均值
            ws.column_dimensions[letter].width=collen*1.0+2	#也就是列宽为最大长度*1.2 可以自己调整
        
        #行和列表头居中
        for c in range(1,ws.max_column + 1):
            for r in range(1,ws.max_row + 1):
                ws.cell(row=r,column=c).alignment = Alignment(horizontal='center', vertical='center')
        wb.save(filename)
